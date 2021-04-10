# encoding=utf-8

import io
import sys
import urllib
# import urllib2
import re
import string
import pymysql
import time
import random
from urllib.request import urlopen
import requests

from django.shortcuts import render
import json
from django.http import HttpResponse,StreamingHttpResponse
from datetime import date,datetime,timedelta
from django.core import serializers
from django.db.models import Avg,Sum,Count



import openpyxl
import datetime
from django.db import connection

# from adviser.views import sql_to_json
# from baseinfo.models import Vip, Serviece,Goods,Cardtype
# from adviser.models import Cardinfo
# from cashier.models import Expvstoll,Expense,Toll


def get_datas(sql,params):
    with connection.cursor() as cursor:
        cursor.execute(sql, params)
        datas=cursor.fetchall()
        cursor.close()
        connection.close()
    return  datas

def get_fields(sql,params):
    with connection.cursor() as cursor:
        cursor.execute(sql, params)
        fields=cursor.description
        cursor.close()
        connection.close()
    return fields

# def get_fields(sql):
#     # 一个传入sql导出字段的函数
#     conn = pms.connect(host='数据库实例地址', user='账号',
#                        passwd='密码', database=库名', port=3306, charset="utf8")
#     cur = conn.cursor()
#     cur.execute(sql)
#     # 获取所需要的字段名称
#     fields = cur.description
#     cur.close()
#     return fields

def get_excel(sql,params,filename,sheettitle):
    # sql=''
    # params=''
    with connection.cursor() as cursor:
        cursor.execute(sql,params)
        datas=cursor.fetchall()
        fields = cursor.description
        cursor.close()
        connection.close()

    # print('fields=',fields)
    print('datas=',datas)

    new = openpyxl.Workbook()
    sheet = new.active
    sheet.title=sheettitle

    for col in range(len(fields)):
        _=sheet.cell(row=1,column=col+1,value=u'%s'%fields[col][0])

    for row in range(len(datas)):
        for col in range(len(fields)):
            _ = sheet.cell(row=row+2,column=col+1,value=u'%s'%datas[row][col])
    newworkbook=new.save(file)
    return newworkbook

def getYesterday():
    # 获取昨天日期的字符串格式的函数
    #获取今天的日期
    today = datetime.date.today()
    #获取一天的日期格式数据
    oneday = datetime.timedelta(days=1)
    #昨天等于今天减去一天
    yesterday = today - oneday
    #获取昨天日期的格式化字符串
    yesterdaystr = yesterday.strftime('%Y-%m-%d')
    #返回昨天的字符串
    return yesterdaystr

def create_email():
    return ''

def send_email():
    return ''

# def main():
#     print(datetime.datetime.now())
#     # my_sql = sql = "SELECT a.id '用户ID',\
#     #        a.gmtCreate '用户注册时间',\
#     #        af.lastLoginTime '最后登录时间',\
#     #        af.totalBuyCount '历史付款子单数',\
#     #        af.paidmountUSD '历史付款金额',\
#     #        af.lastPayTime '用户最后支付时间'\
#     #       FROM table a\
#     #   LEFT JOIN tableb af ON a.id= af.accountId ;"
#     # # 生成数据
#     # my_data = get_datas(my_sql)
#     # # 生成字段名称
#     # my_field = get_fields(my_sql)
#     # 得到昨天的日期
#     # yesterdaystr = getYesterday()
#     # # 文件名称
#     # my_file_name = 'user attribute' + yesterdaystr + '.xlsx'
#     # # 文件路径
#     # file_path = 'D:/work/report/' + my_file_name
#     # # 生成excel
#     # get_excel(my_data, my_field, file_path)
#     #
#     # my_email_from = 'BI部门自动报表机器人'
#     # my_email_to = '运营部'
#     # # 邮件标题
#     # my_email_Subject = 'user' + yesterdaystr
#     # # 邮件正文
#     # my_email_text = "Dear all,\n\t附件为每周数据，请查收！\n\nBI团队 "
#     # # 附件地址
#     # my_annex_path = file_path
#     # # 附件名称
#     # my_annex_name = my_file_name
#     # # 生成邮件
#     # my_msg = create_email(my_email_from, my_email_to, my_email_Subject,
#     #                       my_email_text, my_annex_path, my_annex_name)
#     # my_sender = '阿里云邮箱'
#     # my_password = '我的密码'
#     # my_receiver = [10001 @ qq.com']#接收人邮箱列表
#     #                # 发送邮件
#     #                send_email(my_sender, my_password, my_receiver, my_msg)
#     #                print(datetime.datetime.now())
#     get_excel()
#     if __name__ == "__main__":
#         main();


# select b.storecode, b.vcode, b.vname, b.mtcode, b.viptype, b.viplevel, b.indate, b.birth,
# 	b.ecode, getemplinfo(b.company,b.ecode,'ename') ename, b.ecode2, getemplinfo(b.company,b.ecode2,'ename') ename2,
#     a.class1, a.class2, sum(consumeqty) '消费次数', sum(consumeamount) '消费金额',
# 	sum(leftqty) '疗程剩余次数', sum(timesleftmoney) '疗程剩余金额', sum(amountleftmoney) '储值剩余金额'
# from (
#
# 	select a.vipuuid,
# 		GetAppoptionValue(b.company,'displayclass1',F_GetItemInfobysrvcode(b.srvcode,b.ttype,'displayclass1',b.company)) class1,
# 		GetAppoptionValue(b.company,'displayclass2',F_GetItemInfobysrvcode(b.srvcode,b.ttype,'displayclass2',b.company)) class2,
# 		sum(b.s_qty) consumeqty, sum(b.s_mount *(b.cardratio+b.cashratio)) consumeamount, 0 leftqty, 0 timesleftmoney, 0 amountleftmoney
# 	from expvstoll a, expense b
# 	where 1=1
# 	and a.flag='Y' and b.flag='Y'  and a.valiflag='Y'
# 	and a.company=b.company
# 	and a.company='yiren'
# 	and a.uuid=b.transuuid
# 	and a.vsdate >='20191022'
# 	and a.vsdate <='20191130'
# 	and b.ttype in ('S')
# 	group by a.vipuuid
#
# 	union all
#
# 	select a.vipuuid,
# 		GetAppoptionValue(b.company,'displayclass1',b.displayclass1) class1,
# 		GetAppoptionValue(b.company,'displayclass2', b.displayclass2) class2,
# 		0 consumeqty, 0 consumeamount,
# 		sum(a.leftqty) leftqty, sum(a.leftmoney) timesleftmoney, 0 amountleftmoney
# 	from cardinfo a ,cardtype b
# 	where 1=1
# 	and a.flag='Y' and b.flag='Y'
# 	and a.company=b.company
# 	and a.company='yiren'
# 	and a.cardtype=b.cardtype
# 	and b.suptype in ('20','25')
# 	and b.comptype='times'
# 	group by a.vipuuid
#
# 	union all
#
# 	select a.vipuuid,
# 		GetAppoptionValue(b.company,'displayclass1',b.displayclass1) class1,
# 		GetAppoptionValue(b.company,'displayclass2', b.displayclass2) class2,
# 		0 consumeqty, 0 consumeamount,
# 		0 leftqty,  0 timesleftmoney, sum(a.leftmoney) amountleftmoney
# 	from cardinfo a ,cardtype b
# 	where 1=1
# 	and a.flag='Y' and b.flag='Y'
# 	and a.company=b.company
# 	and a.company='yiren'
# 	and a.cardtype=b.cardtype
# 	and b.suptype ='10'
# 	and b.comptype='amount'
#
# 	group by a.vipuuid
# ) a, vip b
# where 1=1
# and a.vipuuid = b.uuid
# group by b.storecode, b.vcode, b.vname, b.mtcode, b.viptype, b.viplevel, b.indate, b.birth

def get_monthlyreportno1(company,storelist,fromdate,todate):
    host='http://localhost:8080/report/'
    # url = host+'get_monthlyreportno1/?company='+company+'&storecode='+storecode+'&fromdate='+fromdate+'&todate='+todate
    url = host+'get_reportdata_leftmoney/?company='+company+'&storelist='+storelist+'&fromdate='+fromdate+'&todate='+todate
    user_agent = 'Mozilla/4.0 (compatible; MSIE 5.5; Windows NT)'
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
    req =urllib.request.Request(url=url,headers=headers)
    # response = urllib2.urlopen(url)
    # response = urllib.request.urlopen(req)
    html = urllib.request.urlopen(req).read()
    print('html',html)
    return 0

# get_monthlyreportno1('yfy','01','20201001','20201031')

get_monthlyreportno1('yfy','01,02','20201001','20201101')